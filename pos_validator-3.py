#!/usr/bin/env python3
"""
POS Pipeline Validator
=======================
Deterministic, local, field-by-field comparison of Golden vs Generated output
for the three POS output types: POSDATAMJ (EDI 852), POSNLSN (Nielsen/BookScan),
POSDLYM (daily accumulation).

No API keys, no network calls, no LLM/agent judging — every comparison below is
a plain value/format check in Python.

Usage:
    python3 pos_validator.py <golden_path> <generated_path> --out results.json

Handles large files (streams with openpyxl read_only mode; tested comfortably
to 80k+ rows / ~100MB per side).
"""
import sys
import json
import argparse
import collections
from openpyxl import load_workbook

try:
    from python_calamine import CalamineWorkbook
    _HAVE_CALAMINE = True
except ImportError:
    _HAVE_CALAMINE = False

MAX_EXAMPLES = 5  # cap example rows stored per mismatch category, to keep JSON/report small at scale

# Fields that are never compared by default — environment/run-specific by
# nature, not a data-correctness signal (sequence numbers drawn fresh per
# environment/run, and generated-only run identifiers).
DEFAULT_EXCLUDE_FIELDS = {"P0NOID", "PDNOID", "PSNOID", "RunId"}


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def norm(v):
    """Normalize a cell value to a comparable string, preserving leading zeros
    and exact text, but collapsing float-that-is-really-an-int to plain int text."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def numlike(s):
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def values_equal(gv, rv):
    """Return ('exact' | 'soft' | 'diff')."""
    if gv == rv:
        return "exact"
    if numlike(gv) and numlike(rv) and gv != "" and rv != "":
        try:
            if float(gv) == float(rv):
                return "soft"
        except ValueError:
            pass
    return "diff"


def _detect_header(first_row):
    """Shared header heuristic: a header row is all-string, non-empty, and
    doesn't itself look like an EDI/fixed-width data line."""
    vals = [c for c in first_row if c is not None]
    if vals and all(isinstance(c, str) for c in vals) and not str(first_row[0]).startswith(("852", "92", "I3", "94")):
        return True
    return False


def _load_rows_calamine(path, sheet=None):
    wb = CalamineWorkbook.from_path(path)
    sheet_name = sheet or wb.sheet_names[0]
    ws = wb.get_sheet_by_name(sheet_name)
    all_rows = ws.to_python()
    header = None
    rows = []
    for i, r in enumerate(all_rows):
        if i == 0 and _detect_header(r):
            header = list(r)
            continue
        if all(c is None for c in r):
            continue
        rows.append(list(r))
    return header, rows


def _load_rows_openpyxl(path, sheet=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = []
    header = None
    first = True
    for r in ws.iter_rows(values_only=True):
        if first:
            first = False
            if _detect_header(r):
                header = list(r)
                continue
        v = list(r)
        if all(c is None for c in v):
            continue
        rows.append(v)
    wb.close()
    return header, rows


def load_rows(path, sheet=None, max_col_probe=40):
    """Load a workbook's rows as fast as possible. Prefers python-calamine
    (Rust-backed, ~10-15x faster than openpyxl on wide/large sheets — matters
    at 500k+ row scale), falling back to openpyxl if calamine isn't installed
    or fails to open the file for any reason. Both paths are verified to
    produce identical normalized output on every file type this tool handles.
    Returns (header_or_None, rows) — header_or_None is the first row IF it
    looks like text column names, else None (raw single-column fixed-width
    text, e.g. Golden POSDATAMJ/POSNLSN exports)."""
    if _HAVE_CALAMINE:
        try:
            return _load_rows_calamine(path, sheet)
        except Exception:
            pass  # fall through to openpyxl
    return _load_rows_openpyxl(path, sheet)


# ----------------------------------------------------------------------------
# File-type detection
# ----------------------------------------------------------------------------
def detect_type(header, rows):
    if header:
        hset = set(str(h).strip() for h in header if h)
        if "P0AGY" in hset or "P0ISBN" in hset:
            return "POSDLYM"
        if "PSAGY" in hset or "PSITM#" in hset:
            return "POSSUMM"
        if "PDAGY" in hset or "PDITM#" in hset:
            return "POSDTLM"
        if "POSDATA" in hset:
            return "POSDATAMJ"
        if {"RunId", "Seq", "RecordType"} <= hset:
            return "POSNLSN_TABULAR"
    # single-column raw text — sniff first non-empty data line
    for r in rows[:5]:
        v = norm(r[0])
        if v.startswith("852"):
            return "POSDATAMJ"
        if v[:2] in ("92", "I3", "94") or v[:2].isdigit():
            return "POSNLSN_RAW"
    raise ValueError("Could not auto-detect file type from header/content sniff.")


# ----------------------------------------------------------------------------
# POSDATAMJ (EDI 852) parsing
# ----------------------------------------------------------------------------
def parse_posdatamj(lines):
    """Parse raw 852 EDI lines into transaction sets with LIN/ZA/SDQ facts.
    Returns (sets, detail_facts) where detail_facts is keyed
    (item, store, za_qual, za_date) -> qty (int)."""
    sets = []
    cur = None
    curlin = None
    curza = None
    facts = {}
    for idx, raw in enumerate(lines, start=1):
        s = (raw or "").rstrip()
        if len(s) < 3:
            continue
        seg = s[3:6].strip()
        if s.startswith("852000"):
            cur = {"row": idx, "lins": [], "ctt": None, "header_raw": s}
            sets.append(cur)
            curlin = None
        elif seg == "XQ":
            if cur is not None:
                cur["xq_raw"] = s
        elif seg == "LIN":
            curlin = {"row": idx, "qual": s[12:14], "item": s[14:].strip(), "zas": [], "raw": s}
            if cur is not None:
                cur["lins"].append(curlin)
        elif seg == "ZA":
            curza = {"row": idx, "qual": s[6:8], "amount": s[8:18],
                     "dq": s[20:23] if len(s) > 20 else "", "date": s[23:29] if len(s) > 23 else "",
                     "raw": s, "sdqs": []}
            if curlin is not None:
                curlin["zas"].append(curza)
        elif seg == "SDQ":
            uom = s[6:8]
            locq = s[8:10]
            body = s[10:]
            entries = []
            for i in range(0, len(body), 27):
                blk = body[i:i + 27]
                if len(blk) < 27:
                    break
                store = blk[0:17].strip()
                qty = blk[17:27]
                if store:
                    entries.append((store, qty))
                    if curlin is not None and curza is not None:
                        k = (curlin["item"], store, curza["qual"], curza["date"])
                        facts[k] = int(qty)
            if curza is not None:
                curza["sdqs"].append({"row": idx, "uom": uom, "locq": locq, "raw": s, "entries": entries})
        elif seg == "CTT":
            if cur is not None:
                cur["ctt"] = {"row": idx, "count": s[6:12], "raw": s}
    return sets, facts


def get_posdatamj_lines(path):
    """Return the raw 512-byte POSDATA lines regardless of whether the file is
    a single raw-text column (Golden) or a tabular export with FileId/Seq/RunId/
    POSDATA columns (Generated)."""
    header, rows = load_rows(path)
    if header and "POSDATA" in [str(h).strip() for h in header]:
        idx = {str(h).strip(): i for i, h in enumerate(header)}
        col = idx["POSDATA"]
        return [r[col] for r in rows]
    # raw single (or first) column
    return [r[0] for r in rows]


def validate_posdatamj(golden_path, generated_path):
    glines = get_posdatamj_lines(golden_path)
    rlines = get_posdatamj_lines(generated_path)

    gsets, gfacts = parse_posdatamj(glines)
    rsets, rfacts = parse_posdatamj(rlines)

    result = {
        "file_type": "POSDATAMJ",
        "golden_rows": len(glines), "generated_rows": len(rlines),
        "golden_sets": len(gsets), "generated_sets": len(rsets),
        "golden_lins": sum(len(s["lins"]) for s in gsets),
        "generated_lins": sum(len(s["lins"]) for s in rsets),
        "golden_facts": len(gfacts), "generated_facts": len(rfacts),
    }

    missing = []   # generated fact not found in golden
    diffs = []     # found, but qty differs
    matched = 0
    for k, rq in rfacts.items():
        if k not in gfacts:
            missing.append({"key": k, "generated_qty": rq})
        elif gfacts[k] != rq:
            diffs.append({"key": k, "golden_qty": gfacts[k], "generated_qty": rq})
        else:
            matched += 1

    result["facts_matched"] = matched
    result["facts_missing_in_golden"] = len(missing)
    result["facts_qty_mismatch"] = len(diffs)
    result["missing_examples"] = missing[:MAX_EXAMPLES]
    result["diff_examples"] = diffs[:MAX_EXAMPLES]

    # CTT sanity (per set, informational — not compared 1:1 since golden may be a superset run)
    result["golden_ctt"] = [s["ctt"]["count"] for s in gsets if s.get("ctt")]
    result["generated_ctt"] = [s["ctt"]["count"] for s in rsets if s.get("ctt")]
    result["golden_lin_count_matches_ctt"] = [len(s["lins"]) == int(s["ctt"]["count"]) for s in gsets if s.get("ctt")]
    result["generated_lin_count_matches_ctt"] = [len(s["lins"]) == int(s["ctt"]["count"]) for s in rsets if s.get("ctt")]

    return result


# ----------------------------------------------------------------------------
# POSNLSN (Nielsen/BookScan) parsing
# ----------------------------------------------------------------------------
def parse_posnlsn_raw(lines):
    stores = []
    cur = None
    for idx, raw in enumerate(lines, start=1):
        s = (raw or "")
        t = s[:2]
        if t == "92":
            cur = {"row": idx, "chain": s[2:7], "store": s[7:12], "date": s[12:18], "items": [], "trailer": None}
            stores.append(cur)
        elif t == "I3":
            if cur is not None:
                cur["items"].append({"ean": s[2:15], "qty": s[15:20]})
        elif t == "94":
            if cur is not None:
                cur["trailer"] = {"nbr": s[2:7], "tot": s[7:14]}
    return stores


def parse_posnlsn_tabular(header, rows):
    idx = {f: i for i, f in enumerate(header)}
    stores = []
    cur = None
    for r in rows:
        rt = norm(r[idx["RecordType"]])
        if rt == "92":
            cur = {"chain": norm(r[idx["Chain"]]), "store": norm(r[idx["Store"]]),
                   "date": norm(r[idx["PerEndDate"]]), "items": [], "trailer": None}
            stores.append(cur)
        elif rt == "I3":
            if cur is not None:
                cur["items"].append({"ean": norm(r[idx["EANnumber"]]), "qty": norm(r[idx["QtySold"]])})
        elif rt == "94":
            if cur is not None:
                cur["trailer"] = {"nbr": norm(r[idx["NbrRecords"]]), "tot": norm(r[idx["TotQtySold"]])}
    return stores


def load_posnlsn(path):
    header, rows = load_rows(path)
    if header and {"RecordType", "Store"} <= set(str(h).strip() for h in header):
        return parse_posnlsn_tabular(header, rows)
    # Either no header, or a spurious single-cell label row (e.g. 'POSNLSN') —
    # either way this is raw fixed-width text, one record per row.
    lines = [r[0] for r in rows]
    return parse_posnlsn_raw(lines)


def validate_posnlsn(golden_path, generated_path):
    gstores = load_posnlsn(golden_path)
    rstores = load_posnlsn(generated_path)

    gidx = {s["store"].strip(): s for s in gstores}
    ridx = {s["store"].strip(): s for s in rstores}

    result = {
        "file_type": "POSNLSN",
        "golden_stores": len(gstores), "generated_stores": len(rstores),
        "golden_items": sum(len(s["items"]) for s in gstores),
        "generated_items": sum(len(s["items"]) for s in rstores),
    }

    missing_stores = [st for st in ridx if st not in gidx]
    result["stores_missing_in_golden"] = len(missing_stores)
    result["missing_store_examples"] = missing_stores[:MAX_EXAMPLES]

    # item-level compare, only for stores present in both
    item_matched = 0
    item_missing = []
    item_qty_diff = []
    chain_diff = []
    date_diff = []
    trailer_diff = []
    for st, r in ridx.items():
        g = gidx.get(st)
        if not g:
            continue
        if g["chain"].strip() != r["chain"].strip():
            chain_diff.append({"store": st, "golden": g["chain"], "generated": r["chain"]})
        if g["date"].strip() != r["date"].strip():
            date_diff.append({"store": st, "golden": g["date"], "generated": r["date"]})
        gitems = {it["ean"].strip(): it["qty"] for it in g["items"]}
        for it in r["items"]:
            ean = it["ean"].strip()
            if ean not in gitems:
                item_missing.append({"store": st, "ean": ean, "generated_qty": it["qty"]})
            elif gitems[ean] != it["qty"]:
                item_qty_diff.append({"store": st, "ean": ean, "golden_qty": gitems[ean], "generated_qty": it["qty"]})
            else:
                item_matched += 1
        if g["trailer"] and r["trailer"]:
            if g["trailer"] != r["trailer"]:
                trailer_diff.append({"store": st, "golden": g["trailer"], "generated": r["trailer"]})

    result["item_matched"] = item_matched
    result["item_missing_in_golden"] = len(item_missing)
    result["item_qty_mismatch"] = len(item_qty_diff)
    result["chain_mismatch"] = len(chain_diff)
    result["date_mismatch"] = len(date_diff)
    result["trailer_value_mismatch"] = len(trailer_diff)
    result["item_missing_examples"] = item_missing[:MAX_EXAMPLES]
    result["item_qty_diff_examples"] = item_qty_diff[:MAX_EXAMPLES]
    result["trailer_diff_examples"] = trailer_diff[:MAX_EXAMPLES]
    return result


# ----------------------------------------------------------------------------
# POSDLYM (daily accumulation) — tabular, field-by-field
# ----------------------------------------------------------------------------
def validate_tabular(golden_path, generated_path, key_fields, exclude_fields=None, file_type_label="TABULAR"):
    """Generic key-based, field-by-field tabular comparison. Used for POSDLYM,
    POSSUMM, POSDTLM, and any other flat P0/PS/PD-style export."""
    exclude_fields = set(exclude_fields) if exclude_fields is not None else set()
    ghdr, grows = load_rows(golden_path)
    rhdr, rrows = load_rows(generated_path)
    gidx = {f: i for i, f in enumerate(ghdr)}
    ridx = {f: i for i, f in enumerate(rhdr)}

    for f in key_fields:
        if f not in gidx or f not in ridx:
            raise ValueError(f"Validation key field '{f}' not found in both files' headers.")

    def rowkey(row, idx):
        return tuple(norm(row[idx[f]]) for f in key_fields)

    gkey = {}
    gdupe = []
    for row in grows:
        k = rowkey(row, gidx)
        if k in gkey:
            gdupe.append(k)
        gkey[k] = row
    rkey = {}
    rdupe = []
    for row in rrows:
        k = rowkey(row, ridx)
        if k in rkey:
            rdupe.append(k)
        rkey[k] = row

    missing = [k for k in rkey if k not in gkey]          # generated row, no golden counterpart
    extra_golden = [k for k in gkey if k not in rkey]      # golden row, no generated counterpart

    common_fields = [f for f in ghdr if f in ridx and f not in exclude_fields]
    actually_excluded = sorted(set(exclude_fields) & set(ghdr) & set(rhdr))
    field_results = {f: {"exact": 0, "soft": 0, "diff": 0, "soft_examples": [], "diff_examples": []}
                      for f in common_fields}
    rows_all_match = 0
    checked = 0
    for k, r in rkey.items():
        if k not in gkey:
            continue
        g = gkey[k]
        checked += 1
        row_ok = True
        for f in common_fields:
            gv = norm(g[gidx[f]])
            rv = norm(r[ridx[f]])
            verdict = values_equal(gv, rv)
            field_results[f][verdict] += 1
            if verdict == "soft" and len(field_results[f]["soft_examples"]) < MAX_EXAMPLES:
                field_results[f]["soft_examples"].append({"key": k, "golden": gv, "generated": rv})
            if verdict == "diff" and len(field_results[f]["diff_examples"]) < MAX_EXAMPLES:
                field_results[f]["diff_examples"].append({"key": k, "golden": gv, "generated": rv})
            if verdict != "exact":
                row_ok = False
        if row_ok:
            rows_all_match += 1

    result = {
        "file_type": file_type_label,
        "key_fields": list(key_fields),
        "excluded_fields": actually_excluded,
        "golden_rows": len(grows), "generated_rows": len(rrows),
        "golden_keys": len(gkey), "generated_keys": len(rkey),
        "golden_dupes": len(gdupe), "generated_dupes": len(rdupe),
        "rows_missing_in_golden": len(missing), "rows_extra_in_golden": len(extra_golden),
        "missing_examples": missing[:MAX_EXAMPLES], "extra_golden_examples": extra_golden[:MAX_EXAMPLES],
        "rows_checked": checked, "rows_fully_matching": rows_all_match,
        "field_results": field_results,
        "common_fields": common_fields,
    }
    return result


def validate_posdlym(golden_path, generated_path, key_fields=("P0STR", "P0ISBN", "P0ACOD"),
                      exclude_fields=None):
    exclude_fields = exclude_fields if exclude_fields is not None else set(DEFAULT_EXCLUDE_FIELDS)
    return validate_tabular(golden_path, generated_path, key_fields, exclude_fields, "POSDLYM")


def validate_possumm(golden_path, generated_path, key_fields=("PSAGY", "PSCHN", "PSITM#", "PSWEDT"),
                      exclude_fields=None):
    exclude_fields = exclude_fields if exclude_fields is not None else set()
    return validate_tabular(golden_path, generated_path, key_fields, exclude_fields, "POSSUMM")


def validate_posdtlm(golden_path, generated_path, key_fields=("PDAGY", "PDCHN", "PDSTR", "PDITM#", "PDWEDT"),
                      exclude_fields=None):
    exclude_fields = exclude_fields if exclude_fields is not None else set()
    return validate_tabular(golden_path, generated_path, key_fields, exclude_fields, "POSDTLM")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def run_validation(golden_path, generated_path, file_type=None, key=None, exclude=None):
    """Programmatic entry point (used by both the CLI below and server.py).
    key/exclude are comma-separated strings or None (per-type defaults apply)."""
    ftype = file_type
    if ftype is None:
        ghdr, grows = load_rows(golden_path)
        ftype = detect_type(ghdr, grows)
        if ftype in ("POSNLSN_TABULAR", "POSNLSN_RAW"):
            ftype = "POSNLSN"

    if ftype == "POSDATAMJ":
        result = validate_posdatamj(golden_path, generated_path)
    elif ftype == "POSNLSN":
        result = validate_posnlsn(golden_path, generated_path)
    elif ftype == "POSDLYM":
        keys = tuple(k.strip() for k in key.split(",")) if key else ("P0STR", "P0ISBN", "P0ACOD")
        excl = set(x.strip() for x in exclude.split(",") if x.strip()) if exclude is not None else set(DEFAULT_EXCLUDE_FIELDS)
        result = validate_posdlym(golden_path, generated_path, key_fields=keys, exclude_fields=excl)
    elif ftype == "POSSUMM":
        keys = tuple(k.strip() for k in key.split(",")) if key else ("PSAGY", "PSCHN", "PSITM#", "PSWEDT")
        excl = set(x.strip() for x in exclude.split(",") if x.strip()) if exclude is not None else set(DEFAULT_EXCLUDE_FIELDS)
        result = validate_possumm(golden_path, generated_path, key_fields=keys, exclude_fields=excl)
    elif ftype == "POSDTLM":
        keys = tuple(k.strip() for k in key.split(",")) if key else ("PDAGY", "PDCHN", "PDSTR", "PDITM#", "PDWEDT")
        excl = set(x.strip() for x in exclude.split(",") if x.strip()) if exclude is not None else set(DEFAULT_EXCLUDE_FIELDS)
        result = validate_posdtlm(golden_path, generated_path, key_fields=keys, exclude_fields=excl)
    else:
        raise ValueError(f"Unsupported/undetected file type: {ftype}")

    result["_golden_path"] = golden_path
    result["_generated_path"] = generated_path
    return ftype, result


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("golden")
    ap.add_argument("generated")
    ap.add_argument("--type", choices=["POSDATAMJ", "POSNLSN", "POSDLYM", "POSSUMM", "POSDTLM"], default=None,
                     help="Force file type instead of auto-detecting.")
    ap.add_argument("--key", default=None,
                     help="Comma-separated validation key fields (overrides the per-type default).")
    ap.add_argument("--exclude", default="P0NOID,PDNOID,PSNOID,RunId",
                     help="Comma-separated fields to skip entirely in tabular field-by-field comparison "
                          "(default drops known run-specific sequence/id columns; harmless no-op if a "
                          "given file doesn't have that column).")
    ap.add_argument("--out", default="results.json")
    args = ap.parse_args()

    ftype, result = run_validation(args.golden, args.generated, file_type=args.type, key=args.key, exclude=args.exclude)

    with open(args.out, "w") as f:
        json.dump(result, f, indent=2, default=str)
    print(f"Validation complete. Type={ftype}. Results written to {args.out}")


if __name__ == "__main__":
    main()
