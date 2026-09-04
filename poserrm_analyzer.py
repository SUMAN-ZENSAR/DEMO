#!/usr/bin/env python3
"""
POSERRM Analyzer
================
POSERRM is the pipeline's reject-capture table (pos.POSERRM) — one row per
rejected fact, written during phase A4 (Identity Waterfall). This module
reads a POSERRM export and breaks it down by ESRES1 (the reject reason text,
e.g. "ITEM MASTER MISSING", "UPC NOT FOUND", "SCAN MATCH REVERSED"):

  - total row count per ESRES1 category
  - every unique UPC within that category, with its occurrence count
    (and how many distinct stores it showed up at)

This is a single-file analysis, not a Golden-vs-Generated comparison — there
is no "expected" POSERRM to diff against, the file speaks for itself.

Column layout (confirmed from real exports):
ESSAGY ESSCHN ESSSTR ESGLN# ESSITM ESSSKU ESSUPC ESRES1 ESERC1 ESRES2 ESERC2
ESRES3 ESERC3 ESCRDT ESQTYP ESSQTY ESLDTE ESLTIM ESLPRG RunId Seq
"""
import sys
import json
import argparse
import collections

try:
    from python_calamine import CalamineWorkbook
    _HAVE_CALAMINE = True
except ImportError:
    _HAVE_CALAMINE = False

from openpyxl import load_workbook


def norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def _load_rows(path, sheet=None):
    if _HAVE_CALAMINE:
        try:
            wb = CalamineWorkbook.from_path(path)
            ws = wb.get_sheet_by_name(sheet or wb.sheet_names[0])
            all_rows = ws.to_python()
            return all_rows[0], all_rows[1:]
        except Exception:
            pass
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows[0], rows[1:]


def analyze_poserrm(path, sheet=None, max_upcs_per_category=None):
    header, rows = _load_rows(path, sheet)
    idx = {str(h).strip(): i for i, h in enumerate(header)}

    required = ["ESSSTR", "ESSUPC", "ESRES1", "ESERC1", "RunId"]
    missing = [f for f in required if f not in idx]
    if missing:
        raise ValueError(f"POSERRM file is missing expected column(s): {missing}")

    data = [r for r in rows if any(c is not None for c in r)]

    cat_rows = collections.defaultdict(list)
    for r in data:
        cat = norm(r[idx["ESRES1"]])
        cat_rows[cat].append(r)

    def field(r, name, default=""):
        return norm(r[idx[name]]) if name in idx else default

    categories = []
    for cat, rlist in cat_rows.items():
        upc_counter = collections.Counter()
        upc_stores = collections.defaultdict(set)
        store_set = set()
        qty_signed_total = 0
        code = field(rlist[0], "ESERC1")
        for r in rlist:
            upc = field(r, "ESSUPC")
            store = field(r, "ESSSTR")
            upc_counter[upc] += 1
            upc_stores[upc].add(store)
            store_set.add(store)
            try:
                qty_signed_total += int(field(r, "ESSQTY") or 0)
            except ValueError:
                pass

        upc_breakdown = [
            {"upc": upc, "count": cnt, "stores_affected": len(upc_stores[upc])}
            for upc, cnt in upc_counter.most_common(max_upcs_per_category)
        ]
        categories.append({
            "category": cat,
            "code": code,
            "total_count": len(rlist),
            "unique_upcs": len(upc_counter),
            "stores_affected": len(store_set),
            "net_qty": qty_signed_total,
            "upc_breakdown": upc_breakdown,
        })

    categories.sort(key=lambda c: c["total_count"], reverse=True)

    run_ids = sorted(set(field(r, "RunId") for r in data), key=lambda x: (len(x), x))
    stores_all = sorted(set(field(r, "ESSSTR") for r in data), key=lambda x: (len(x), x))
    chains = sorted(set(field(r, "ESSCHN") for r in data))
    agencies = sorted(set(field(r, "ESSAGY") for r in data))

    return {
        "total_rows": len(data),
        "distinct_run_ids": run_ids,
        "chains": chains,
        "agencies": agencies,
        "distinct_stores_affected": len(stores_all),
        "categories": categories,
        "_source_file": str(path),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("poserrm_file")
    ap.add_argument("--out", default="poserrm_results.json")
    ap.add_argument("--max-upcs", type=int, default=None, help="Cap unique-UPC rows shown per category (default: all)")
    args = ap.parse_args()

    result = analyze_poserrm(args.poserrm_file, max_upcs_per_category=args.max_upcs)
    with open(args.out, "w") as f:
        json.dump(result, f, indent=2, default=str)
    print(f"POSERRM analysis complete. {result['total_rows']} rows, {len(result['categories'])} categories. Written to {args.out}")


if __name__ == "__main__":
    main()
